import os
import csv
import sys
import glob
import time

# pip install pandas openpyxl (in conda environment)
import pandas as pd
import gzip
import re
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.cell.rich_text import CellRichText, TextBlock, InlineFont

# run in bash:
# my_env=pacvar_output_format
# script_dir="/kyukon/home/gent/422/vsc42287/BIT11/sync/HPC_scripts/output_formatting"
# output_dir="/kyukon/home/gent/422/vsc42287/BIT11/local_output/pacvar_test_karyotype"
# samplesheet="/kyukon/home/gent/422/vsc42287/BIT11/sync/HPC_scripts/run_nf_pacvar/samplesheet_karyotype.csv"
# conda run -n "$my_env" python3 "$script_dir/generate_puretarget_xlsx_report.py" "$output_dir" "$samplesheet"

def read_samplesheet(samplesheet_path):
    with open(samplesheet_path, newline='') as csvfile:
        return list(csv.DictReader(csvfile))

def color_motifs_in_sequence(seq, motifs, header):
    motif_colors = ['2B8FCC','8E60D6', 'E47941','EF4D8E','F1D22E','1AA760','E64B35','F2E998']   
    color_map = {}
    motifs = sorted(motifs, key=len, reverse=True)
    for i in range(len(motifs)):
        motif = motifs[i]
        color = motif_colors[i % len(motif_colors)]  # wrap around if more motifs than colors
        color_map[motif] = color
    colored_seq = CellRichText()    

    i = 0
    while i < len(seq):
        matched = False
        for motif, color in color_map.items():
            if seq[i:i+len(motif)] == motif:
                colored_seq.append(TextBlock(InlineFont(color=color, b=True, rFont="Consolas"), motif))
                i += len(motif)
                matched = True
                break
        if not matched:
            if header:
                colored_seq.append(TextBlock(InlineFont(color="000000"), seq[i]))
            else: 
                colored_seq.append(TextBlock(InlineFont(color="FF0000", b=True, u="single", rFont="Consolas"), seq[i]))
            i += 1
    return colored_seq

def write_report(sample_name, output_dir):  

    os.makedirs(os.path.join(output_dir, "output_report"), exist_ok=True)

    filepath = os.path.join(output_dir, "trgt")
    vcf_file = os.path.join(filepath, f"{sample_name}.vcf.gz")
    output_excel = os.path.join(output_dir, "output_report", f"{sample_name}_PureTarget_report.xlsx")

    # Extract TRGT version from VCF
    trgt_version = ""
    with gzip.open(vcf_file, 'rt') as f:
        for line in f:
            if "##trgtVersion=" in line:
                trgt_version = re.search(r'trgtVersion=(\S+)', line).group(1)
                trgt_version = trgt_version.split('-')[0]
                break  # stop loop once found

    # Parse VCF
    vcf_data = []
    with gzip.open(vcf_file, 'rt') as f:
        for line in f:
            if line.startswith("#"):
                continue
            parts = line.strip().split('\t')
            chrom, pos, _, ref, alt, _, _, info, fmt, smpl = parts
            info_dict = dict([x.split('=') for x in info.split(';')])
            fmt_keys = fmt.split(':')
            smpl_values = smpl.split(':')
            fmt_dict = dict(zip(fmt_keys, smpl_values))
            vcf_data.append({
                "CHROM": chrom,
                "POS": pos,
                "REF": ref,
                "ALT": alt,
                "TRID": info_dict.get("TRID",""),
                "MOTIFS": info_dict.get("MOTIFS",""),
                "GT": fmt_dict.get("GT",""),
                "AL": fmt_dict.get("AL",""),
                "ALLR": fmt_dict.get("ALLR",""),
                "SD": fmt_dict.get("SD",""),
                "MC": fmt_dict.get("MC",""),
                "MS": fmt_dict.get("MS",""),
                "AP": fmt_dict.get("AP",""),
                "AM": fmt_dict.get("AM","")
    })

    vcf_df = pd.DataFrame(vcf_data)
    all_trids = sorted(vcf_df['TRID'].unique())

    # Create Excel workbook
    wb = Workbook()
    ws_menu = wb.active
    ws_menu.title = "Menu"

    bold_font = Font(name='Calibri', bold=True)
    grey_fill = PatternFill(start_color='FFE0E0E0', end_color='FFE0E0E0', fill_type='solid')

    ws_menu["A1"] = "PureTarget report"
    ws_menu["A1"].font = Font(size=18, bold=True, color="FFE31B92")
    ws_menu["A2"] = f"{sample_name}"
    ws_menu["A2"].font = bold_font
    ws_menu["A3"] = f"TRGT v{trgt_version}"
    ws_menu["A3"].font = bold_font
    ws_menu["A5"] = "Select repeat"
    ws_menu["A5"].font = Font(bold=True, color="FFE31B92")

    # Place TRIDs in hidden helper column for dropdown list
    ws_menu["BA1"] = "TRID_list"
    for i, trid in enumerate(all_trids, start=2):
        ws_menu[f"BA{i}"] = trid

    # Create dropdown validation
    dv = DataValidation(type="list", formula1=f"=BA2:BA{len(all_trids)+1}")
    ws_menu.add_data_validation(dv)
    ws_menu["B5"] = all_trids[0] # default
    dv.add(ws_menu["B5"])

    # Set a formula that sets hyperlink target dynamically
    ws_menu["A7"].font = Font(bold=True, color="FF0070C0")
    ws_menu["A7"] = '=HYPERLINK("#\'"&B5&"\'!A1", "▶ Click here to open selected repeat")'

    # Formatting of 'menu' sheet
    for col in ws_menu.columns:
        ws_menu.column_dimensions[col[0].column_letter].width = 15

    # Now create sheets for each TRID
    for trid in all_trids:
        ws_trid = wb.create_sheet(title=trid[:31])

        # Write summary for repeat(s) of interest
        for idx, row in vcf_df.iterrows():

            # Skip rows that are NOT for this TRID
            if row['TRID'] != trid:
                continue

            ws_trid.append(["Repeat", row['TRID']])
            ws_trid.cell(row=ws_trid.max_row, column=1).font = bold_font

            ws_trid.append(["Position", f"{row['CHROM']}:{row['POS']}"])
            ws_trid.cell(row=ws_trid.max_row, column=1).font = bold_font

            motifs_colored = color_motifs_in_sequence(seq=row['MOTIFS'], motifs=row['MOTIFS'].split(','), header=True)
            ws_trid.append(["Expected motifs", motifs_colored])
            ws_trid.cell(row=ws_trid.max_row, column=1).font = bold_font

            for row_idx in range(ws_trid.max_row-2,ws_trid.max_row+1):
                for col_idx in range(1, 300):
                    ws_trid.cell(row=row_idx, column=col_idx).fill = grey_fill
            ws_trid.append([])
            ws_trid.append(["Allele", "Length", "Length range", "Reads spanning allele", "Motif count", "Motif span", "Purity", "Mean methylation", "Sequence"])
            for cell in ws_trid[ws_trid.max_row]:
                cell.font = bold_font

            # Determine alleles based on GT
            gt = row.get("GT", "")
            ref = row["REF"]
            alt_alleles = row["ALT"].split(",") if row["ALT"] else []

            if "/" in gt:
                allele_indices = gt.split("/")
            else:
                # "0" or "1" for chrX
                allele_indices = [gt]       

            # Convert to integers (ignore '.')
            allele_indices = [int(a) for a in allele_indices if a.isdigit()]   

            # Helper to safely split list fields
            def safe_split(value):
                return value.split(",") if value else []

            al_list  = safe_split(row.get("AL",""))
            allr_list = safe_split(row.get("ALLR",""))
            sd_list  = safe_split(row.get("SD",""))
            mc_list  = safe_split(row.get("MC",""))
            ms_list  = safe_split(row.get("MS",""))
            ap_list  = safe_split(row.get("AP",""))
            am_list  = safe_split(row.get("AM",""))

            # Build a mapping
                # 0:REF
                # 1:ALT_1
                # 2:ALT_2
            allele_map = {0: ref}
            for i, alt in enumerate(alt_alleles):
                allele_map[i+1] = alt

            # Loop through alleles in GT and write one row per allele
            for idx, allele_index in enumerate(allele_indices):
                allele_seq = allele_map.get(allele_index, "")
                allele_seq_colored = color_motifs_in_sequence(seq=allele_seq, motifs=row['MOTIFS'].split(','), header=False)
                allele_label = f"Allele {idx+1}"            
                row_cells = [
                    allele_label,
                    al_list[idx],
                    allr_list[idx],
                    sd_list[idx],
                    mc_list[idx],
                    ms_list[idx],
                    ap_list[idx],
                    am_list[idx],
                    allele_seq_colored
                ]
                ws_trid.append(row_cells)
            ws_trid.append([])
            ws_trid.append([])
        
            # Final formatting of 'report' sheet
            for col in ws_trid.columns:
                ws_trid.column_dimensions[col[0].column_letter].width = 15

    # # For each repeat, print images on separate sheet
    # for trid in sorted(vcf_df['TRID'].unique()):
    #     # Excel sheet name max 31 chars
    #     sheet_name = trid[:31]  
    #     ws_trid = wb.create_sheet(title=sheet_name)
    #     motif_png = os.path.join(filepath, f"{sample_name}_{trid}_motifs.png")
    #     meth_png  = os.path.join(filepath, f"{sample_name}_{trid}_meth.png")

    #     # Define starting location
    #     row_pos = 1
    #     col_pos = 1

    #     for label, cell, png_path  in [("Motifs", "A3", motif_png), ("Methylation", "M3", meth_png)]:
    #         if os.path.exists(png_path):
    #             # Insert a label
    #             ws_trid.cell(row=row_pos, column=col_pos, value=f"{label} {trid}")
    #             ws_trid.cell(row=row_pos, column=col_pos).font = bold_font
    #             col_pos += 12

    #             # Insert image and reset starting location
    #             img = Image(png_path)
    #             ws_trid.add_image(img, cell)

    # Save workbook
    wb.save(output_excel)
    print(f"Report saved to {output_excel}")

def main():

    start = time.time()

    if len(sys.argv) != 3:
        print(f"\nUsage: python generate_puretarget_xlsx_report.py <output_dir> <samplesheet>")
        sys.exit(1)

    output_dir = sys.argv[1]
    samplesheet = sys.argv[2]

    samples = read_samplesheet(samplesheet)
    for sample in samples:
        sample_name = sample['sample']
        write_report(sample_name=sample_name, output_dir=output_dir)

    end = time.time()
    print(f"\nElapsed time: {end - start}")

if __name__ == "__main__":
    main()