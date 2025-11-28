import os
import csv
import sys
import glob
import time

import pandas as pd
import gzip
import re
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, PatternFill
from openpyxl.cell.rich_text import CellRichText, TextBlock, InlineFont

def color_motifs_in_sequence(seq, motifs):
    motif_colors = ['2B8FCC','8E60D6', 'E47941','EF4D8E','F1D22E','1AA760','E64B35','F2E998']   
    color_map = {}
    motifs = sorted(motifs, key=len, reverse=True)
    for i in range(len(motifs)):
        motif = motifs[i]
        color = motif_colors[i % len(motif_colors)]  # wrap around if more motifs than colors
        color_map[motif] = color
    colored_seq = CellRichText()    

    i = 0
    while i < len(seq):
        matched = False
        for motif, color in color_map.items():
            if seq[i:i+len(motif)] == motif:
                colored_seq.append(TextBlock(InlineFont(color=color, b=True, rFont="Consolas"), motif))
                i += len(motif)
                matched = True
                break
        if not matched:
            colored_seq.append(TextBlock(InlineFont(color="FF0000", b=True, u="single", rFont="Consolas"), seq[i]))
            i += 1
    return colored_seq



    rt = CellRichText()

    for char, color in seq:        # however you generate the blocks
        rt.append(TextBlock(text=char, font=InlineFont(color=color)))

    cell = ws.cell(row=1, column=1)
    



def write_report(seq, motifs):  
      
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "report"
    seq_colored = color_motifs_in_sequence(seq,motifs.split(','))
    cell = ws.cell(row=1, column=1)
    cell.value = seq_colored

    # Save workbook
    output_excel = "test_color_motifs.xlsx"
    wb.save(output_excel)
    print(f"Report saved to {output_excel}")

def main():
    seq = "GGAAAATAAAATGAAATGAAAGAAATGAAAAAAT"
    motifs = "AAAAT,GAAAT"
    write_report(seq, motifs)

if __name__ == "__main__":
    main()